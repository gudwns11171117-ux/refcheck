# -*- coding: utf-8 -*-
"""검증 결과를 엑셀(xlsx)로 내보낸다."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FILL = {
    "확인됨": "C6EFCE",
    "검토 필요": "FFEB9C",
    "확인 불가": "FFC7CE",
    "검증 제외": "E7E6E6",
    "오류": "E7E6E6",
}
STATUS_ORDER = ["확인 불가", "검토 필요", "검증 제외", "오류", "확인됨"]


def to_xlsx(results: list[dict], source_name: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "참고문헌 검증"
    head = ["번호", "판정", "주의", "자료유형", "참고문헌(원문)", "추출 제목", "추출 연도", "확인된 제목", "확인된 연도",
            "확인된 저자", "게재지/발행처", "DOI", "제목 유사도(%)", "조회한 곳", "대표 링크", "RISS", "기타 링크", "비고"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in results:
        ref = r.get("ref") or {}
        best = r.get("best") or {}
        links = r.get("links") or []
        primary = next((l["url"] for l in links if l.get("kind") == "primary"), "")
        riss_l = next((l["url"] for l in links if "RISS" in l.get("label", "")), "")
        others = "\n".join(f"{l['label']}: {l['url']}" for l in links if l["url"] not in (primary, riss_l))
        row = [
            ref.get("index"), r.get("status_label"), ", ".join(r.get("flags") or []), r.get("kind_label"),
            ref.get("raw"), ref.get("title"), ref.get("year"),
            best.get("title"), best.get("year"), ", ".join(best.get("authors") or [])[:200], best.get("container"),
            best.get("doi"), best.get("title_sim") if best else None, ", ".join(r.get("sources") or []),
            primary, riss_l, others, r.get("note"),
        ]
        ws.append(row)
        cell = ws.cell(row=ws.max_row, column=2)
        color = FILL.get(r.get("status_label", ""), None)
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
        if r.get("flags"):
            ws.cell(row=ws.max_row, column=3).font = Font(color="CF222E", bold=True)
        for col in (15, 16):
            c = ws.cell(row=ws.max_row, column=col)
            if c.value:
                c.hyperlink = c.value
                c.font = Font(color="0563C1", underline="single")
    widths = [6, 10, 22, 12, 58, 38, 8, 38, 8, 28, 24, 24, 10, 20, 38, 38, 46, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("요약")
    ws2.append(["원본 파일", source_name or "(붙여넣은 텍스트)"])
    ws2.append(["전체 건수", len(results)])
    for label in STATUS_ORDER:
        ws2.append([label, sum(1 for r in results if r.get("status_label") == label)])
    need = sum(1 for r in results if r.get("status_label") != "확인됨")
    ws2.append(["직접 확인이 필요한 건수", need])
    ws2.append([])
    ws2.append(["주의 표시", "건수"])
    counts: dict[str, int] = {}
    for r in results:
        for f in r.get("flags") or []:
            counts[f] = counts.get(f, 0) + 1
    for k, v in sorted(counts.items(), key=lambda x: -x[1]):
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 40
    for c in (ws2["A1"], ws2["A2"]):
        c.font = Font(bold=True)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
